#encoding=utf-8
import urllib2
import urllib
import cookielib
import json
import uniout
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from BeautifulSoup import BeautifulSoup
import os

idx = 0
txt = ''
order ='desc'
oderby = 'd.id'
total = 0
totalPage = 0
auditStatus = ''
status = ''

baseUrl = "http://adexchange.cn.miaozhen.com" 
getDeals = baseUrl+"/deal/list" #
viewOne = baseUrl+'/deal/view?id='
baseNum = 500
maxPages = 5

class JSONHelper(object):
    """docstring for JSONHelper"""
    def __init__(self,jc):
        super(JSONHelper, self).__init__()
        self.jsonContent = json.loads(jc[jc.find('{') : ]) 
    def getAllUNAudiIds(self):
        jdata = self.jsonContent 
        total = jdata['total']
        mapper = jdata["tableData"]
        idList = list()
        for o in mapper:
            idList.append(str(o['Id'])+'_'+str(o['AuditStatus'] ))
        return idList
    def auditStatus(self):
        jdata = self.jsonContent
        total = jdata['AuditStatus']
        return total

class xlsxHelper(object):
    """docstring for outputRes"""
    def __init__(self, allDetails,filename):
        super(xlsxHelper, self).__init__()
        self.allDetails = allDetails
        self.filename = filename

    def exportAsXls(self):
        firstC = ['DealID','Deal名称*','CPM价格（分/千次曝光）*','对应上游媒体','上游DealID','总预算（元）','每日预算（元）','投放日期*','优先级','投放地区','投放设备类型','投放操作系统','DSP*','广告位*','广告主*','流量分布','审核状态','链接']
        tarwb = Workbook()
        dest_filename = self.filename
        ws1 = tarwb.active
        ws1.title = "all"
        #生成列名
        fcn = 1
        for fc in firstC:
            ws1.cell(column=fcn, row=1).value = str(fc)
            fcn = fcn + 1
        #填写数据
        r = 2 
        for row in self.allDetails:
            c = 1
            for cell in row:
                ws1.cell(column=c, row=r).value = cell
                c = c +1
            r = r +1
        tarwb.save(filename = dest_filename)

class Details(object):
    """docstring for getAllDetails"""
    def __init__(self, html):
        super(Details, self).__init__()
        self.html = html
        
    def getOneRow(self):
        soup = BeautifulSoup(self.html )
        tables = soup.findChildren('table')
        my_table = tables[0]
        rows = my_table.findChildren(['th', 'tr'])
        res = list()
        for row in rows:
            cells = row.findChildren('td')
            listM = cells[::-2] 
            for cell in listM:
                temp = cell.string
                if temp:
                    res.append(temp.replace(' ','').replace('\n','').strip().encode('utf-8'))
                else:
                    res.append('empty')
                    # print 'empty'
        return res     


def initCredential():
    #获得一个cookieJar实例,
    cj = cookielib.CookieJar()
    #定义一个opener 作为后面带着cookie访问的对象
    opener=urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))
    #用户名和，密码，明文，注意安全哦
    user = ''
    password = ''
    login_page = "http://adexchange.cn.miaozhen.com/j_spring_security_check"
    try:
        formdata = {'j_username':user, 'j_password':password }
        #form data
        ldata = urllib.urlencode(formdata)
        opener.addheaders = [('User-agent','Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)')]
        opener.open("http://adexchange.cn.miaozhen.com/j_spring_security_check",ldata)
        return opener
    except Exception,e:
        print str(e)

def getJC(opener,url,pageNum,lines): 
    dealList = {'current':pageNum,'lines':lines,'idx':0,'order':order,'orderBy':oderby,'txt':txt,'total':0,'totalPage':0,'auditStatus':'','status':''}
    dldata = urllib.urlencode(dealList);
    request = urllib2.Request(url,dldata)
    response = opener.open(request)
    content = response.read() 
    return content

def getDetails(opener,viewOne,dId):
    url = viewOne+str(dId)
    request = urllib2.Request(url)
    response = opener.open(request)
    content = response.read()
    return content

opener = initCredential()
for i in range(1,maxPages):
    filename = 'dealsFile_'+str(i)+".xlsx"
    allDetails = list ()
    dealsContent = getJC(opener,getDeals,i,baseNum)
    jh =  JSONHelper(dealsContent)
    temp = jh.getAllUNAudiIds()
    if temp :
        for Id in temp:
            auditStatus = Id[-1]
            html = getDetails(opener,viewOne,Id[0:-2])
            link = viewOne + Id[0:-2]
            hd = Details(html)
            tempList = hd.getOneRow()
            tempList.append(auditStatus)
            tempList.append(link)
            allDetails.append(tempList)
    else:
        #程序结束
        break
    gXls = xlsxHelper(allDetails,filename)
    gXls.exportAsXls()


 
# print temp
# print jh.getTotal()